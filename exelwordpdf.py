import openpyxl
import os

# check directory to where example document is in is in
os.chdir('/Users/kj/Desktop/Certificates/Automate the Boring Stuff')

# create a workbook object by opening the exel file
workbook - openpyxl.load_workbook('example.xlsx')

# to get sheet object if you know the sheets name
sheet = workbook.get_sheet_by_name('Sheet1')

# if you dont know the name of the sheets - you can get list of string values of sheets
workbook.get_sheet_names()

# to get the cells from the sheets
cell = sheet['A1']

print(cell.value)

# make it a string
print(str(cell.value))

# get another cell value
cell = sheet['B1'].value

# you can also get cells using cell - not that rows and columns start at 1 not 0
cell = sheet.cell(row=1, column=2)

print(cell)

# loop over multiple cells
for i in range(1, 8):
    # get the first 7 rows of the second column
    print(i, sheet.cell(row=1, column=2).value)



