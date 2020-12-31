import openpyxl, os

# create a new workbook
wb = openpyxl.Workbook()

# get teh sheet object
sheet = wb.get_sheet_by_name('Sheet')

# update data in a cell
sheet['A1'] = 42
sheet['A2'] = 'Hello'

# change directory you want to save into
os.chdir('/Users/kj/Desktop/Certificates/Automate the Boring Stuff')

# save the document
wb.save('example.xlsx')

# add new worksheet to file
sheet2 = wb.create_sheet()

# notice there are now 2 sheets in excel file
wb.get_sheet_names()

# change the name of a sheet
sheet2.title = 'stocks'

wb.save('example2.xlsx')

# if we want to add new sheet to part instead of appending at the end - give it index and title
wb.create_sheet(index=0, title='stockprofits')